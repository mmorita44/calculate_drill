import random
import traceback
import openpyxl


class Drill:

    def __init__(self, file):
        '''コンストラクタ

        Args:
            file (str): 書き込み先のファイル名
        Returns:
            None
        '''
        self.row_max = 27
        self.cloumn_max = 20
        self.file = file

    def write(self, first_min, first_max, second_min, second_max):
        '''計算式の書き込み

        Args:
            first_min (int): 1項目の最小値
            first_max (int): 1項目の最大値
            first_max (int): 2項目の最小値
            first_max (int): 3項目の最大値
        Returns:
            None
        '''
        try:
            book = openpyxl.load_workbook(self.file)
            sheet = book['Sheet1']
            records = []
            for first in range(first_min, first_max):
                for second in range(second_min, second_max):
                    records.append([first, second])
            random.shuffle(records)

            row_pos = 1
            column_pos = 1
            for record in records:
                if row_pos >= self.row_max:
                    row_pos = 1
                    column_pos += 5

                if column_pos >= self.cloumn_max:
                    break

                sheet.cell(row=row_pos, column=column_pos, value=record[0])
                sheet.cell(row=row_pos, column=column_pos + 2, value=record[1])
                row_pos += 2

            book.save(self.file)
            return 0
        except Exception:
            print(traceback.format_exc())
            return 1
